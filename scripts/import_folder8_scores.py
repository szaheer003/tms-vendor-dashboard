"""
Import evaluator xlsx files from Folder 8 (glob *.xlsx) into:
  - src/data/evaluatorScores.json (raw matrix + qualitative + confidence + proceed)
  - updates src/data/portfolio.json scorecard + vendor composites + radar

Re-run whenever new files appear in Folder 8. Filenames may change; vendor is detected from
"Vendor N" or vendor name in the filename.
"""
from __future__ import annotations

import json
import re
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
FOLDER8 = ROOT / "Folder 8"
SRC_DATA = ROOT / "src" / "data"
PUBLIC_DATA = ROOT / "public" / "data"

# Column header matching (header lowercased; all phrases must appear)
SUB_HEADER_RULES: list[tuple[str, tuple[str, ...]]] = [
    ("5.1", ("reference & case study",)),
    ("5.2", ("operational depth",)),
    ("5.3", ("contracting posture",)),
    ("2.1", ("sla commitment",)),
    ("2.2", ("language, coverage",)),
    ("2.3", ("delivery model credibility",)),
    ("3.1", ("tech governance compliance",)),
    ("3.2", ("migration plan specificity",)),
    ("3.3", ("ai overlay innovation",)),
    ("4.1", ("4.1 wave", "wave composition")),
    ("4.2", ("4.2 sequencing", "sequencing, speed")),
    ("4.3", ("4.3 workforce", "workforce transition")),
    ("1.1", ("1.1 synergy", "synergy alignment")),
    ("1.2", ("1.2 efficiency", "efficiency assumptions")),
    ("1.3", ("1.3 investments", "investments, transparency")),
]

# Each question: one or more phrase groups. A column matches if ALL phrases in ANY group appear
# in the header (Google Forms / Sheets export wording varies).
QUAL_RULES: list[tuple[str, tuple[tuple[str, ...], ...]]] = [
    ("Q3", (("like most",), ("liked most",))),
    (
        "Q4",
        (
            ("need to be true", "approach"),
            ("need to be true", "vendor"),
            ("need to be true", "proposal"),
            ("what would need to be true",),
            ("for this vendor", "need to be true"),
            ("vendor's approach", "need to be true"),
        ),
    ),
    ("Q5", (("biggest risk",), ("major risk",))),
    (
        "Q6",
        (
            ("drill into", "workshop"),
            ("drill into", "next"),
            ("questions would you", "drill"),
            ("questions do you have", "drill"),
            ("want to drill",),
        ),
    ),
    ("Q7", (("emulate", "all vendors should"), ("anything this vendor demonstrated", "emulate"))),
]


def vendor_id_from_filename(name: str) -> str | None:
    n = name.lower()
    if "cognizant" in n or "vendor 1" in n or "vendor_1" in n:
        return "cognizant"
    if "genpact" in n or "vendor 2" in n:
        return "genpact"
    if "exl" in n and "excel" not in n or "vendor 3" in n:
        return "exl"
    if "sutherland" in n or "vendor 4" in n:
        return "sutherland"
    if "ubiquity" in n or "vendor 5" in n:
        return "ubiquity"
    if "ibm" in n or "vendor 6" in n:
        return "ibm"
    return None


def find_col_map(ws) -> dict[str, int]:
    headers = []
    for c in range(1, ws.max_column + 1):
        v = ws.cell(1, c).value
        headers.append((c, (str(v).lower() if v is not None else "")))
    col_by_sub: dict[str, int] = {}
    for sub_id, phrases in SUB_HEADER_RULES:
        for c, h in headers:
            if all(p in h for p in phrases):
                col_by_sub[sub_id] = c
                break
    return col_by_sub


def find_qual_cols(ws) -> dict[str, int]:
    headers = [(c, str(ws.cell(1, c).value or "").lower()) for c in range(1, ws.max_column + 1)]
    out: dict[str, int] = {}

    def pick_col(alternatives: tuple[tuple[str, ...], ...]) -> int | None:
        for phrases in alternatives:
            for c, h in headers:
                if all(p in h for p in phrases):
                    return c
        return None

    for qid, alternatives in QUAL_RULES:
        col = pick_col(alternatives)
        if col is not None:
            out[qid] = col

    # Fallbacks when wording still does not match any rule
    if "Q4" not in out:
        for c, h in headers:
            if "need to be true" in h:
                out["Q4"] = c
                break
    if "Q6" not in out:
        for c, h in headers:
            if "drill" in h and ("workshop" in h or "next" in h):
                out["Q6"] = c
                break

    return out


def find_conf_proceed(ws) -> tuple[int | None, int | None]:
    headers = [(c, str(ws.cell(1, c).value or "").lower()) for c in range(1, ws.max_column + 1)]
    conf = proc = None
    for c, h in headers:
        if "confidence level" in h and conf is None:
            conf = c
        if "proceed" in h and "next" in h and proc is None:
            proc = c
    return conf, proc


def parse_score(val) -> float | None:
    if val is None or val == "":
        return None
    if isinstance(val, (int, float)) and not isinstance(val, bool):
        return float(val)
    s = str(val).strip()
    if not s:
        return None
    try:
        return float(s.replace(",", ""))
    except ValueError:
        return None


def parse_proceed(val) -> bool | None:
    if val is None:
        return None
    s = str(val).strip().lower()
    if s in ("yes", "y", "true", "1"):
        return True
    if s in ("no", "n", "false", "0"):
        return False
    return None


def mean(xs: list[float]) -> float | None:
    return sum(xs) / len(xs) if xs else None


PILLAR_META: list[tuple[tuple[str, str, str], str]] = [
    (("1.1", "1.2", "1.3"), "Commercial Attractiveness (22.5%)"),
    (("2.1", "2.2", "2.3"), "Operational Excellence & Delivery (22.5%)"),
    (("3.1", "3.2", "3.3"), "Technology & AI (22.5%)"),
    (("4.1", "4.2", "4.3"), "Client & Workforce Migration (22.5%)"),
    (("5.1", "5.2", "5.3"), "Partnership Readiness (10%)"),
]

WEIGHTS = {
    "commercial": 0.225,
    "operations": 0.225,
    "technology": 0.225,
    "migration": 0.225,
    "partnership": 0.1,
}

SUB_TO_PILLAR_KEY: dict[str, str] = {}
for subs, label in PILLAR_META:
    key = (
        "commercial"
        if "Commercial" in label
        else "operations"
        if "Operational" in label
        else "technology"
        if "Technology" in label
        else "migration"
        if "Workforce Migration" in label
        else "partnership"
    )
    for sid in subs:
        SUB_TO_PILLAR_KEY[sid] = key


def main():
    files = sorted(FOLDER8.glob("*.xlsx"))
    if not files:
        print("No xlsx in Folder 8", file=__import__("sys").stderr)
        return 1

    all_vendors: dict[str, dict] = {}
    max_evaluators = 0

    for fp in files:
        vid = vendor_id_from_filename(fp.name)
        if not vid:
            print("skip (unknown vendor):", fp.name)
            continue
        wb = openpyxl.load_workbook(fp, data_only=True)
        ws = wb.active
        cmap = find_col_map(ws)
        qmap = find_qual_cols(ws)
        conf_c, proc_c = find_conf_proceed(ws)
        req = [r[0] for r in SUB_HEADER_RULES]
        for r in req:
            if r not in cmap:
                print(f"WARN {fp.name}: missing column for {r}")

        eval_rows: list[dict] = []
        for r in range(2, ws.max_row + 1):
            scores_row: dict[str, float | None] = {}
            for sid, col in cmap.items():
                scores_row[sid] = parse_score(ws.cell(r, col).value)
            if all(v is None for v in scores_row.values()):
                continue
            qual = {
                qk: (str(ws.cell(r, qmap[qk]).value or "").strip() if qk in qmap else "")
                for qk, _ in QUAL_RULES
            }
            conf = str(ws.cell(r, conf_c).value or "").strip() if conf_c else ""
            proc = parse_proceed(ws.cell(r, proc_c).value) if proc_c else None
            name = str(ws.cell(r, 7).value or "").strip() or f"Evaluator {len(eval_rows)+1}"
            eval_rows.append(
                {
                    "displayName": name,
                    "scores": scores_row,
                    "qualitative": qual,
                    "confidence": conf or None,
                    "proceed": proc,
                }
            )
        all_vendors[vid] = {"file": fp.name, "evaluators": eval_rows}
        max_evaluators = max(max_evaluators, len(eval_rows))

    # Build matrix ev1..evN (N = max_evaluators) for JSON flat structure
    vendor_ids = ["cognizant", "genpact", "exl", "sutherland", "ubiquity", "ibm"]
    sub_ids = [r[0] for r in SUB_HEADER_RULES]

    scores_matrix: dict[str, dict[str, dict[str, float | None]]] = {}
    qualitative_matrix: dict[str, dict[str, dict[str, str]]] = {}
    confidence_matrix: dict[str, dict[str, str | None]] = {}
    proceed_matrix: dict[str, dict[str, bool | None]] = {}

    for vid in vendor_ids:
        scores_matrix[vid] = {}
        qualitative_matrix[vid] = {}
        confidence_matrix[vid] = {}
        proceed_matrix[vid] = {}
        evs = all_vendors.get(vid, {}).get("evaluators", [])
        for i in range(max_evaluators):
            ekey = f"ev{i+1}"
            scores_matrix[vid][ekey] = {s: None for s in sub_ids}
            qualitative_matrix[vid][ekey] = {q: "" for q, _ in QUAL_RULES}
            confidence_matrix[vid][ekey] = None
            proceed_matrix[vid][ekey] = None
        for i, er in enumerate(evs):
            ekey = f"ev{i+1}"
            for sid, val in er["scores"].items():
                if sid in scores_matrix[vid][ekey]:
                    scores_matrix[vid][ekey][sid] = val
            for qk, qv in er.get("qualitative", {}).items():
                if qk in qualitative_matrix[vid][ekey]:
                    qualitative_matrix[vid][ekey][qk] = qv
            confidence_matrix[vid][ekey] = er.get("confidence")
            proceed_matrix[vid][ekey] = er.get("proceed")

    out = {
        "importedAt": datetime.now(timezone.utc).isoformat(),
        "source": "Folder 8 — Workshop 1 evaluator Google Form exports (xlsx)",
        "evaluatorSlotCount": max_evaluators,
        "vendorsPresent": [v for v in vendor_ids if v in all_vendors],
        "scores": scores_matrix,
        "qualitative": qualitative_matrix,
        "confidence": confidence_matrix,
        "proceed": proceed_matrix,
        "filesImported": {v: all_vendors[v]["file"] for v in all_vendors},
    }

    SRC_DATA.mkdir(parents=True, exist_ok=True)
    PUBLIC_DATA.mkdir(parents=True, exist_ok=True)
    (SRC_DATA / "evaluatorScores.json").write_text(json.dumps(out, indent=2), encoding="utf-8")
    (PUBLIC_DATA / "evaluatorScores.json").write_text(json.dumps(out, indent=2), encoding="utf-8")

    # Aggregates for portfolio scorecard
    dim_avgs: dict[str, dict[str, float | None]] = defaultdict(dict)
    for sid in sub_ids:
        for vid in vendor_ids:
            vals = [
                scores_matrix[vid][f"ev{i+1}"][sid]
                for i in range(max_evaluators)
                if scores_matrix[vid][f"ev{i+1}"][sid] is not None
            ]
            nums = [float(x) for x in vals if x is not None]
            dim_avgs[sid][vid] = round(mean(nums), 3) if nums else None

    pillar_avgs: dict[str, dict[str, float | None]] = {k: {} for k in WEIGHTS}
    for vid in vendor_ids:
        for pkey in WEIGHTS:
            subs = [s for s, pk in SUB_TO_PILLAR_KEY.items() if pk == pkey]
            nums = [dim_avgs[s][vid] for s in subs if dim_avgs[s][vid] is not None]
            pillar_avgs[pkey][vid] = round(mean([float(x) for x in nums]), 3) if nums else None

    composite: dict[str, float | None] = {}
    for vid in vendor_ids:
        acc = 0.0
        tw = 0.0
        for pkey, w in WEIGHTS.items():
            pav = pillar_avgs[pkey][vid]
            if pav is not None:
                acc += w * pav
                tw += w
        composite[vid] = round(acc / tw, 3) if tw > 0 else None

    portfolio = json.loads((SRC_DATA / "portfolio.json").read_text(encoding="utf-8"))
    portfolio["scorecard"]["source"] = out["source"] + f" · Imported {out['importedAt'][:10]}."
    label_by_sub = {}
    for d in portfolio["scorecard"]["dimensions"]:
        label_by_sub[d["id"]] = d["label"]

    new_dims = []
    for subs, pillar_label in PILLAR_META:
        for sid in subs:
            new_dims.append(
                {
                    "id": sid,
                    "pillar": pillar_label,
                    "label": label_by_sub.get(sid, sid),
                    "scores": {vid: dim_avgs[sid].get(vid) for vid in vendor_ids},
                }
            )
    portfolio["scorecard"]["dimensions"] = new_dims
    portfolio["scorecard"]["composite"] = composite

    for v in portfolio["vendors"]:
        if v["id"] in composite:
            v["composite"] = composite[v["id"]]

    # Radar: scale 0-10 from 1-9 scores -> (val-1)/8*10
    def scale_radar(x: float | None) -> float:
        if x is None:
            return 0.0
        return max(0.0, min(10.0, (float(x) - 1.0) / 8.0 * 10.0))

    field_sum = defaultdict(float)
    field_n = defaultdict(int)
    for vid in vendor_ids:
        for pk in WEIGHTS:
            pv = pillar_avgs[pk][vid]
            if pv is not None:
                field_sum[pk] += scale_radar(pv)
                field_n[pk] += 1
    portfolio["radar"]["fieldAverage"] = {
        k: round(field_sum[k] / field_n[k], 3) if field_n[k] else 0.0 for k in WEIGHTS
    }
    for rv in portfolio["radar"]["vendors"]:
        vid = rv["vendorId"]
        rv["pillars"] = {k: scale_radar(pillar_avgs[k].get(vid)) for k in WEIGHTS}

    (SRC_DATA / "portfolio.json").write_text(json.dumps(portfolio, indent=2), encoding="utf-8")
    (PUBLIC_DATA / "portfolio.json").write_text(json.dumps(portfolio, indent=2), encoding="utf-8")

    # Sync scorecard.json shell with same dimensions (for any code reading it)
    sc = json.loads((SRC_DATA / "scorecard.json").read_text(encoding="utf-8"))
    sc["source"] = portfolio["scorecard"]["source"]
    sc["dimensions"] = new_dims
    sc["composite"] = composite
    (SRC_DATA / "scorecard.json").write_text(json.dumps(sc, indent=2), encoding="utf-8")
    (PUBLIC_DATA / "scorecard.json").write_text(json.dumps(sc, indent=2), encoding="utf-8")

    print("Imported", len(all_vendors), "vendors,", max_evaluators, "evaluator slots")
    print("Wrote evaluatorScores.json, portfolio.json, scorecard.json (src + public data)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
